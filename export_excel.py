import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from config import REPORTS_PATH

class ExcelExporter:
    
    def __init__(self):
        if not os.path.exists(REPORTS_PATH):
            os.makedirs(REPORTS_PATH)
    
    def exportar_inventario(self, productos):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventario"
            
            header_fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            headers = ["ID", "Nombre", "Descripción", "Cantidad", "Precio Unitario", 
                      "Valor Total", "Proveedor", "Fecha Registro", "Última Actualización"]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            for row_num, producto in enumerate(productos, 2):
                datos_fila = [
                    producto.get('id'),
                    producto.get('nombre', 'N/A'),
                    producto.get('descripcion', ''),
                    producto.get('cantidad', 0),
                    float(producto.get('precio_unitario', 0)),
                    float(producto.get('cantidad', 0)) * float(producto.get('precio_unitario', 0)),
                    producto.get('proveedor', 'N/A'),
                    producto.get('fecha_registro', ''),
                    producto.get('ultima_actualizacion', ''),
                ]
                
                for col_num, valor in enumerate(datos_fila, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = valor
                    cell.border = border
                    
                    if col_num in [1, 4]:
                        cell.alignment = Alignment(horizontal="center")
                    elif col_num in [5, 6]:
                        cell.number_format = '$#,##0.00'
                        cell.alignment = Alignment(horizontal="right")
            
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 18
            ws.column_dimensions['I'].width = 18
            
            ws.freeze_panes = "A2"
            
            filename = f"{REPORTS_PATH}Inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            return True, filename
        
        except Exception as err:
            return False, f"Error al exportar inventario: {str(err)}"
    
    def exportar_movimientos(self, movimientos, productos_dict):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Movimientos"
            
            header_fill = PatternFill(start_color="2a9d8f", end_color="2a9d8f", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            headers = ["ID Movimiento", "Producto", "Tipo", "Cantidad", "Fecha", "Descripción"]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            for row_num, mov in enumerate(movimientos, 2):
                id_producto = mov.get('id_producto')
                producto_nombre = (productos_dict.get(id_producto, {}).get('nombre', 'N/A') 
                                 if productos_dict else 'N/A')
                
                datos_fila = [
                    mov.get('id'),
                    producto_nombre,
                    mov.get('tipo_movimiento', 'N/A'),
                    mov.get('cantidad', 0),
                    mov.get('fecha', ''),
                    mov.get('descripcion', ''),
                ]
                
                for col_num, valor in enumerate(datos_fila, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = valor
                    cell.border = border
                    
                    if col_num in [1, 4]:
                        cell.alignment = Alignment(horizontal="center")
            
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 30
            
            ws.freeze_panes = "A2"
            
            filename = f"{REPORTS_PATH}Movimientos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            return True, filename
        
        except Exception as err:
            return False, f"Error al exportar movimientos: {str(err)}"
    
    def exportar_completo(self, productos, movimientos, productos_dict):
        try:
            wb = openpyxl.Workbook()
            
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            header_fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            ws_inv = wb.create_sheet("Inventario", 0)
            headers_inv = ["ID", "Nombre", "Descripción", "Cantidad", "Precio Unitario", 
                          "Valor Total", "Proveedor", "Fecha Registro", "Última Actualización"]
            
            for col_num, header in enumerate(headers_inv, 1):
                cell = ws_inv.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            for row_num, producto in enumerate(productos, 2):
                datos_fila = [
                    producto.get('id'),
                    producto.get('nombre', 'N/A'),
                    producto.get('descripcion', ''),
                    producto.get('cantidad', 0),
                    float(producto.get('precio_unitario', 0)),
                    float(producto.get('cantidad', 0)) * float(producto.get('precio_unitario', 0)),
                    producto.get('proveedor', 'N/A'),
                    producto.get('fecha_registro', ''),
                    producto.get('ultima_actualizacion', ''),
                ]
                
                for col_num, valor in enumerate(datos_fila, 1):
                    cell = ws_inv.cell(row=row_num, column=col_num)
                    cell.value = valor
                    cell.border = border
                    
                    if col_num in [1, 4]:
                        cell.alignment = Alignment(horizontal="center")
                    elif col_num in [5, 6]:
                        cell.number_format = '$#,##0.00'
                        cell.alignment = Alignment(horizontal="right")
            
            ws_inv.column_dimensions['A'].width = 8
            ws_inv.column_dimensions['B'].width = 25
            ws_inv.column_dimensions['C'].width = 30
            ws_inv.column_dimensions['D'].width = 12
            ws_inv.column_dimensions['E'].width = 15
            ws_inv.column_dimensions['F'].width = 15
            ws_inv.column_dimensions['G'].width = 20
            ws_inv.column_dimensions['H'].width = 18
            ws_inv.column_dimensions['I'].width = 18
            ws_inv.freeze_panes = "A2"
            
            ws_mov = wb.create_sheet("Movimientos", 1)
            headers_mov = ["ID Movimiento", "Producto", "Tipo", "Cantidad", "Fecha", "Descripción"]
            
            for col_num, header in enumerate(headers_mov, 1):
                cell = ws_mov.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            for row_num, mov in enumerate(movimientos, 2):
                id_producto = mov.get('id_producto')
                producto_nombre = (productos_dict.get(id_producto, {}).get('nombre', 'N/A') 
                                 if productos_dict else 'N/A')
                
                datos_fila = [
                    mov.get('id'),
                    producto_nombre,
                    mov.get('tipo_movimiento', 'N/A'),
                    mov.get('cantidad', 0),
                    mov.get('fecha', ''),
                    mov.get('descripcion', ''),
                ]
                
                for col_num, valor in enumerate(datos_fila, 1):
                    cell = ws_mov.cell(row=row_num, column=col_num)
                    cell.value = valor
                    cell.border = border
                    
                    if col_num in [1, 4]:
                        cell.alignment = Alignment(horizontal="center")
            
            ws_mov.column_dimensions['A'].width = 15
            ws_mov.column_dimensions['B'].width = 25
            ws_mov.column_dimensions['C'].width = 15
            ws_mov.column_dimensions['D'].width = 12
            ws_mov.column_dimensions['E'].width = 18
            ws_mov.column_dimensions['F'].width = 30
            ws_mov.freeze_panes = "A2"
            
            ws_resumen = wb.create_sheet("Resumen", 2)
            ws_resumen['A1'] = "RESUMEN DE INVENTARIO"
            ws_resumen['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws_resumen['A1'].fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
            ws_resumen.merge_cells('A1:B1')
            ws_resumen['A1'].alignment = Alignment(horizontal="center", vertical="center")
            
            ws_resumen['A2'] = "Fecha de Generación:"
            ws_resumen['B2'] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            
            ws_resumen['A4'] = "Estadísticas Generales:"
            ws_resumen['A4'].font = Font(bold=True, size=11)
            
            ws_resumen['A5'] = "Total de Productos:"
            ws_resumen['B5'] = len(productos)
            
            total_stock = sum(p.get('cantidad', 0) for p in productos)
            ws_resumen['A6'] = "Stock Total:"
            ws_resumen['B6'] = total_stock
            
            valor_total = sum(float(p.get('cantidad', 0)) * float(p.get('precio_unitario', 0)) 
                            for p in productos)
            ws_resumen['A7'] = "Valor Total del Inventario:"
            ws_resumen['B7'] = valor_total
            ws_resumen['B7'].number_format = '$#,##0.00'
            
            bajo_stock = sum(1 for p in productos if p.get('cantidad', 0) < 10)
            ws_resumen['A8'] = "Productos con Stock Bajo (<10):"
            ws_resumen['B8'] = bajo_stock
            
            total_movimientos = len(movimientos)
            ws_resumen['A9'] = "Total de Movimientos:"
            ws_resumen['B9'] = total_movimientos
            
            ws_resumen.column_dimensions['A'].width = 30
            ws_resumen.column_dimensions['B'].width = 20
            
            filename = f"{REPORTS_PATH}Inventario_Completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            return True, filename
        
        except Exception as err:
            return False, f"Error al exportar datos completos: {str(err)}"
